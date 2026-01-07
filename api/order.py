from http.server import BaseHTTPRequestHandler
import json
import os
import requests
import sys
from datetime import datetime, timedelta, timezone
import io, tempfile, json, os, requests
import csv

sys.path.append(os.path.dirname(__file__))

try:
    from supabase_init import supabase
    from health import log_error
except ImportError as e:
    print(f"Import error: {e}")

# Добавляем импорт для работы с Excel
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.pagebreak import Break
except ImportError as e:
    print(f"⚠️ Openpyxl import error: {e}")
    # Создаем заглушки для совместимости
    openpyxl = None

class Handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS, GET, PUT, DELETE')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Telegram-Id, Is-Admin, User-Id')
        self.end_headers()
    
    def do_GET(self):
        try:
            print(f"📥 GET запрос на путь: {self.path}")
            
            # Добавляем обработку экспорта заказов
            if self.path == '/api/order/export' or self.path.startswith('/api/order/export'):
                print("🔄 Вызов handle_export_orders из do_GET")
                return self.handle_export_orders()
            
            user_id = self.headers.get('User-Id', '')
            is_admin = self.headers.get('Is-Admin', 'false') == 'true'
            
            if is_admin:
                orders_response = supabase.table("orders").select("*").execute()
                statuses_response = supabase.table("order_statuses").select("*").execute()
                
                orders = orders_response.data
                statuses = statuses_response.data
                
                status_map = {status['id']: status for status in statuses}
                
                for order in orders:
                    status_info = status_map.get(order['status_id'])
                    if status_info:
                        order['status_name'] = status_info['name']
                        order['status_color'] = status_info['color']
            else:
                response = supabase.table("orders").select("*").eq("user_id", user_id).execute()
                orders = response.data
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            self.wfile.write(json.dumps(orders).encode('utf-8'))
            
        except Exception as e:
            log_error("order_GET", e, self.headers.get('User-Id', ''), "Failed to fetch orders")
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            response = {'success': False, 'error': str(e)}
            self.wfile.write(json.dumps(response).encode('utf-8'))
    
    def do_PUT(self):
        try:
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            order_data = json.loads(post_data)
            
            order_id = order_data.get('order_id')
            status_id = order_data.get('status_id')
            
            if not order_id:
                raise ValueError("Order ID is required")
            
            update_data = {'status_id': status_id}
            
            if status_id == 5:
                order_response = supabase.table("orders").select("total_amount, discount_amount").eq("id", order_id).execute()
                if order_response.data:
                    order = order_response.data[0]
                    profit = order['total_amount'] - (order['discount_amount'] or 0)
                    update_data['profit'] = profit
            
            response = supabase.table("orders").update(update_data).eq("id", order_id).execute()
            
            if status_id:
                self.send_order_notification(order_id, status_id)
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            response_data = {'success': True, 'message': 'Order updated successfully'}
            self.wfile.write(json.dumps(response_data).encode('utf-8'))
            
        except Exception as e:
            log_error("order_PUT", e, self.headers.get('User-Id', ''), f"Order ID: {order_data.get('order_id')}")
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            response = {'success': False, 'error': str(e)}
            self.wfile.write(json.dumps(response).encode('utf-8'))
    
    def do_POST(self):
        try:
            print(f"📥 POST запрос на путь: {self.path}")
            
            # Убираем обработку экспорта из POST, теперь она в GET
            if self.path == '/api/order/export' or self.path.startswith('/api/order/export'):
                print("❌ Экспорт должен вызываться через GET метод")
                self.send_response(405)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                response = {'success': False, 'error': 'Используйте GET метод для экспорта'}
                self.wfile.write(json.dumps(response).encode('utf-8'))
                return
                
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            order_data = json.loads(post_data)
            
            user_id = str(order_data['user']['id'])
            
            db_success = self.save_order_to_db(order_data)
            
            if db_success:
                delivery_option = order_data.get('delivery_option', 'pickup')
                delivery_address = order_data.get('delivery_address', '')
                discount_amount = order_data.get('discount_amount', 0)
                promocode_id = order_data.get('promocode_id')
                
                admin_success = self.send_admin_notification(order_data, delivery_option, delivery_address, discount_amount)
                
                # Отправляем уведомление клиенту
                customer_success = self.send_customer_notification(order_data)
                
                if promocode_id:
                    self.update_promocode_usage(promocode_id)
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            response = {'success': True, 'message': 'Order processed successfully', 'db_success': db_success}
            self.wfile.write(json.dumps(response).encode('utf-8'))
            
        except Exception as e:
            log_error("order_POST", e, order_data.get('user', {}).get('id', ''), "Failed to create order")
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            response = {'success': False, 'error': str(e)}
            self.wfile.write(json.dumps(response).encode('utf-8'))
    
    def do_DELETE(self):
        try:
            path_parts = self.path.split('/')
            order_id = path_parts[-1] if path_parts[-1] else path_parts[-2]
            
            if not order_id.isdigit():
                self.send_response(400)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                response = {'success': False, 'error': 'Invalid order ID'}
                self.wfile.write(json.dumps(response).encode('utf-8'))
                return
            
            response = supabase.table("orders").delete().eq("id", int(order_id)).execute()
            
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            response_data = {'success': True}
            self.wfile.write(json.dumps(response_data).encode('utf-8'))
            
        except Exception as e:
            log_error("order_DELETE", e, self.headers.get('User-Id', ''), f"Order ID: {order_id}")
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            response = {'success': False, 'error': str(e)}
            self.wfile.write(json.dumps(response).encode('utf-8'))
    
    def get_moscow_time(self):
        """Получение текущего времени по Москве (UTC+3)"""
        # Текущее время в UTC
        utc_now = datetime.now(timezone.utc)
        # Добавляем 3 часа для московского времени
        moscow_offset = timedelta(hours=3)
        moscow_time = utc_now + moscow_offset
        return moscow_time
    
    def convert_utc_to_moscow(self, utc_dt_str):
        """Конвертирует UTC время из базы данных в московское время"""
        try:
            # Парсим UTC время из строки
            utc_dt = datetime.fromisoformat(utc_dt_str.replace('Z', '+00:00'))
            # Добавляем 3 часа для московского времени
            moscow_offset = timedelta(hours=3)
            moscow_dt = utc_dt + moscow_offset
            return moscow_dt
        except Exception as e:
            print(f"⚠️ Ошибка преобразования времени: {e}")
            return None
    
    def send_admin_notification(self, order_data, delivery_option, delivery_address, discount_amount):
        try:
            bot_token = os.environ.get('BOT_TOKEN')
            
            admins_response = supabase.table("admins").select("telegram_id").eq("is_active", True).execute()
            admin_chat_ids = [admin['telegram_id'] for admin in admins_response.data]
            
            if not bot_token or not admin_chat_ids:
                log_error("order_notification", "Missing BOT_TOKEN or no active admins", order_data['user']['id'], "Admin notification failed")
                return False
            
            clean_phone = order_data['phone'].replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
            telegram_link = f"tg://openmessage?user_id={order_data['user']['id']}"
            
            delivery_info = "🚚 Доставка" if delivery_option == "delivery" else "🏪 Самовывоз"
            if delivery_option == "delivery" and delivery_address:
                delivery_info += f"\n📍 Адрес: {delivery_address}"
            else:
                settings_response = supabase.table("shop_settings").select("value").eq("key", "contacts").execute()
                if settings_response.data:
                    contacts = settings_response.data[0]['value']
                    pickup_address = contacts.get('address', 'Ярославль, ул. Цветочная, 15')
                    delivery_info += f"\n📍 Адрес самовывоза: {pickup_address}"
            
            items_text = "\n".join([
                f"• {item['name']} - {item['quantity']} шт. × {item['price']} ₽ = {item['total']} ₽" 
                for item in order_data['items']
            ])
            
            cart_total = order_data['total']
            delivery_cost = 0
            free_delivery_min = 3000
            
            if delivery_option == "delivery":
                settings_response = supabase.table("shop_settings").select("value").eq("key", "delivery_price").execute()
                if settings_response.data:
                    delivery_price = settings_response.data[0]['value'].get('value', 200)
                    free_delivery_min_response = supabase.table("shop_settings").select("value").eq("key", "free_delivery_min").execute()
                    if free_delivery_min_response.data:
                        free_delivery_min = free_delivery_min_response.data[0]['value'].get('value', 3000)
                    
                    delivery_cost = 0 if cart_total >= free_delivery_min else delivery_price
            
            total_with_delivery = cart_total + delivery_cost - discount_amount
            
            discount_text = f"🎫 Скидка по промокоду: -{discount_amount} ₽\n" if discount_amount > 0 else ""
            
            message = f"""🎉 *НОВЫЙ ЗАКАЗ!*

👤 *Информация о клиенте:*
🆔 ID: `{order_data['user']['id']}`
📛 Имя: {order_data['user']['first_name']}
👤 Юзернейм: @{order_data['user']['username']}
📞 Телефон: `{clean_phone}`

{delivery_info}

🛍️ *Состав заказа:*
{items_text}

💵 *Сумма заказа:* {cart_total} ₽
🚚 *Доставка:* {f'{delivery_cost} ₽' if delivery_cost > 0 else 'Бесплатно'} {f'(бесплатно от {free_delivery_min} ₽)' if delivery_cost > 0 else ''}
{discount_text}💎 *Итого к оплате:* {total_with_delivery} ₽

📋 *Комментарий:* {order_data.get('comment', 'Нет комментария')}

🕐 *Время заказа:* {order_data['time']}

💬 *Связаться с клиентом:*
[📱 Написать в Telegram]({telegram_link})"""
            
            success_count = 0
            for admin_chat_id in admin_chat_ids:
                url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                payload = {
                    'chat_id': admin_chat_id,
                    'text': message,
                    'parse_mode': 'Markdown',
                    'disable_web_page_preview': True,
                    'reply_markup': {
                        'inline_keyboard': [[
                            {'text': '📱 Написать клиенту', 'url': telegram_link}
                        ]]
                    }
                }
                
                response = requests.post(url, json=payload, timeout=10)
                if response.status_code == 200:
                    success_count += 1
            
            return success_count > 0
            
        except Exception as e:
            log_error("admin_notification", e, order_data['user']['id'], "Failed to send admin notification")
            return False

    def send_customer_notification(self, order_data):
        try:
            bot_token = os.environ.get('BOT_TOKEN')
            user_id = order_data['user']['id']
            
            if not bot_token:
                log_error("customer_notification", "Missing BOT_TOKEN", user_id, "Failed to send customer notification")
                return False
            
            # Форматируем товары
            items_text = "\n".join([
                f"• {item['name']} - {item['quantity']} шт." 
                for item in order_data['items']
            ])
            
            # Форматируем телефон для красивого отображения
            phone = order_data['phone']
            # Убираем все нецифровые символы
            digits = ''.join(filter(str.isdigit, phone))
            
            if len(digits) >= 11:
                # Если номер начинается с 7, 8 или +7
                if digits.startswith('7') or digits.startswith('8'):
                    if digits.startswith('8'):
                        digits = '7' + digits[1:]
                    formatted_phone = f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
                else:
                    formatted_phone = phone
            else:
                formatted_phone = phone
            
            # Используем время из заказа или текущее время
            order_time = order_data.get('time', self.get_moscow_time().strftime('%d.%m.%Y, %H:%M:%S'))
            
            # Формируем сообщение
            message = f"""✅ Ваш заказ принят!

🛍 Состав заказа:
{items_text}

💵 Сумма заказа: {order_data['total']} ₽

📞 Ваш телефон: {formatted_phone}

⏱ Время заказа: {order_time}

Мы свяжемся с вами в ближайшее время для подтверждения заказа и уточнения деталей доставки.

Спасибо за ваш заказ! 💐"""
            
            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            payload = {
                'chat_id': user_id,
                'text': message,
                'parse_mode': 'Markdown'
            }
            
            response = requests.post(url, json=payload, timeout=10)
            return response.status_code == 200
            
        except Exception as e:
            log_error("customer_notification", e, user_id, "Failed to send customer notification")
            return False

    def save_order_to_db(self, order_data):
        try:
            clean_phone = order_data['phone'].replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
            
            cart_total = order_data['total']
            delivery_option = order_data.get('delivery_option', 'pickup')
            delivery_address = order_data.get('delivery_address', '')
            promocode_id = order_data.get('promocode_id')
            discount_amount = order_data.get('discount_amount', 0)
            
            delivery_cost = 0
            free_delivery_min = 3000
            
            if delivery_option == "delivery":
                try:
                    settings_response = supabase.table("shop_settings").select("value").eq("key", "delivery_price").execute()
                    if settings_response.data:
                        delivery_price = settings_response.data[0]['value'].get('value', 200)
                        free_delivery_min_response = supabase.table("shop_settings").select("value").eq("key", "free_delivery_min").execute()
                        if free_delivery_min_response.data:
                            free_delivery_min = free_delivery_min_response.data[0]['value'].get('value', 3000)
                        
                        delivery_cost = 0 if cart_total >= free_delivery_min else delivery_price
                except Exception as e:
                    print(f"⚠️ Delivery settings error: {e}")
            
            final_amount = cart_total + delivery_cost - discount_amount
            
            order_record = {
                "user_id": str(order_data['user']['id']),
                "user_name": order_data['user']['first_name'],
                "user_username": order_data['user'].get('username', ''),
                "phone": clean_phone,
                "comment": order_data.get('comment', ''),
                "delivery_option": delivery_option,
                "delivery_address": delivery_address,
                "items": order_data['items'],
                "total_amount": cart_total,
                "discount_amount": discount_amount,
                "final_amount": final_amount,
                "promocode_id": promocode_id,
                "status_id": 1,
                "profit": 0
            }
            
            result = supabase.table("orders").insert(order_record).execute()
            
            if result.data:
                return True
            else:
                return False
                
        except Exception as e:
            print(f"💥 Error saving order to database: {e}")
            return False

    def handle_export_orders(self):
        try:
            print(f"🔄 Начало обработки экспорта заказов через GET")
            
            bot_token = os.environ.get('BOT_TOKEN')
            user_id = self.headers.get('Telegram-Id', '')
            is_admin = self.headers.get('Is-Admin', 'false') == 'true'
            
            print(f"📊 Параметры запроса: bot_token={'установлен' if bot_token else 'отсутствует'}, user_id={user_id}, is_admin={is_admin}")
            
            if not bot_token:
                print("❌ Отсутствует BOT_TOKEN")
                self.send_response(500)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                response_data = {'success': False, 'error': 'Отсутствует BOT_TOKEN'}
                self.wfile.write(json.dumps(response_data).encode('utf-8'))
                return
            
            if not user_id:
                print("❌ Отсутствует Telegram-Id")
                self.send_response(400)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                response_data = {'success': False, 'error': 'Требуется авторизация'}
                self.wfile.write(json.dumps(response_data).encode('utf-8'))
                return
            
            if not is_admin:
                print("❌ Пользователь не является администратором")
                self.send_response(403)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                response_data = {'success': False, 'error': 'Требуются права администратора'}
                self.wfile.write(json.dumps(response_data).encode('utf-8'))
                return
            
            print("📋 Запрашиваем заказы из базы данных...")
            orders_response = supabase.table("orders").select("*").order('created_at', desc=True).execute()
            
            if not orders_response.data:
                print("⚠️ Нет данных для экспорта")
                self.send_response(200)
                self.send_header('Content-type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                response_data = {'success': True, 'message': 'Нет данных для экспорта', 'data': []}
                self.wfile.write(json.dumps(response_data).encode('utf-8'))
                return
            
            print(f"✅ Найдено {len(orders_response.data)} заказов")
            
            # Проверяем, установлен ли openpyxl
            if openpyxl is None:
                print("⚠️ Библиотека openpyxl не установлена, используем CSV")
                return self.export_to_csv(orders_response.data, bot_token, user_id)
            
            # Создаем Excel файл с улучшенным форматированием
            return self.export_to_excel(orders_response.data, bot_token, user_id)
                
        except Exception as e:
            error_msg = str(e)
            print(f"💥 Критическая ошибка в экспорте: {error_msg}")
            log_error("export_orders", e, self.headers.get('Telegram-Id', ''), "Ошибка экспорта")
            
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            response_data = {'success': False, 'error': f'Ошибка сервера: {error_msg}'}
            self.wfile.write(json.dumps(response_data).encode('utf-8'))

    def export_to_excel(self, orders, bot_token, user_id):
        """Создание и отправка Excel файла с профессиональным дизайном"""
        try:
            print("📊 Создаем профессиональный Excel отчет...")
            
            # Создаем рабочую книгу
            wb = Workbook()
            
            # Удаляем стандартный лист
            if len(wb.sheetnames) > 0:
                std_sheet = wb[wb.sheetnames[0]]
                wb.remove(std_sheet)
            
            # ===== ЛИСТ 1: ДЕТАЛИЗАЦИЯ ЗАКАЗОВ =====
            ws1 = wb.create_sheet(title="📋 Детализация заказов")
            
            # Определяем стили
            title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            cell_font = Font(name='Calibri', size=10)
            cell_alignment = Alignment(horizontal='center', vertical='center')
            
            money_font = Font(name='Calibri', size=10, bold=True, color='1F4E78')
            status_font = Font(name='Calibri', size=10, bold=True)
            product_font = Font(name='Calibri', size=10, color='2E4053')
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            
            # Добавляем заголовок отчета
            ws1.merge_cells('A1:M1')
            title_cell = ws1.cell(row=1, column=1, value=f"📊 Отчет по заказам - АРТФЛОРА")
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Добавляем подзаголовок с датой (МОСКОВСКОЕ ВРЕМЯ)
            ws1.merge_cells('A2:M2')
            moscow_time = self.get_moscow_time()
            subtitle_cell = ws1.cell(row=2, column=1, value=f"Сформирован: {moscow_time.strftime('%d.%m.%Y %H:%M')} (МСК)")
            subtitle_cell.font = Font(name='Calibri', size=10, italic=True, color='7F7F7F')
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Заголовки столбцов
            headers = [
                '№', 'ID заказа', 'Дата создания', 'Клиент', 'Телефон',
                'Состав заказа', 'Кол-во товаров', 'Сумма (₽)', 'Скидка (₽)', 
                'Итог (₽)', 'Способ', 'Статус', 'Примечание'
            ]
            
            # Записываем заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws1.cell(row=4, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Карта статусов с улучшенными цветами
            status_names = {
                1: ('🆕 Новый', 'FF6B6B'),
                2: ('✅ Подтвержден', 'FFA726'),
                3: ('📦 Собирается', '8E44AD'),
                4: ('🚚 В пути', '3498DB'),
                5: ('🎉 Доставлен', '27AE60'),
                6: ('❌ Отменен', '95A5A6')
            }
            
            row_num = 5
            summary_data = {
                'total_orders': len(orders),
                'total_amount': 0,
                'total_discount': 0,
                'total_final': 0,
                'total_products': 0,
                'delivery_count': 0,
                'pickup_count': 0,
                'status_counts': {status_id: 0 for status_id in status_names.keys()}
            }
            
            # Словарь для статистики по товарам
            product_stats = {}
            
            # Заполняем данные
            for idx, order in enumerate(orders, 1):
                # Форматируем дату (МОСКОВСКОЕ ВРЕМЯ)
                order_time = ''
                if order.get('created_at'):
                    try:
                        # Преобразуем UTC время в московское
                        moscow_dt = self.convert_utc_to_moscow(order['created_at'])
                        if moscow_dt:
                            order_time = moscow_dt.strftime('%d.%m.%Y\n%H:%M')
                        else:
                            order_time = str(order['created_at'])
                    except:
                        order_time = str(order['created_at'])
                
                # Получаем статус
                status_info = status_names.get(order['status_id'], ('❓ Неизвестен', 'CCCCCC'))
                status_text, status_color = status_info
                summary_data['status_counts'][order['status_id']] += 1
                
                # Форматируем телефон
                phone = order['phone']
                if len(phone) >= 10:
                    formatted_phone = f"+7 ({phone[1:4]}) {phone[4:7]}-{phone[7:9]}-{phone[9:11]}"
                else:
                    formatted_phone = phone
                
                # Получаем товары
                items = []
                try:
                    if isinstance(order['items'], str):
                        items = json.loads(order['items'])
                    else:
                        items = order['items']
                except:
                    items = []
                
                # Формируем строку с составом заказа
                items_text_parts = []
                total_items_in_order = 0
                
                for item in items:
                    item_name = item.get('name', 'Неизвестный товар')
                    item_quantity = item.get('quantity', 0)
                    item_price = item.get('price', 0)
                    item_total = item.get('total', 0)
                    
                    items_text_parts.append(f"• {item_name} × {item_quantity} шт. = {item_total} ₽")
                    total_items_in_order += item_quantity
                    
                    # Собираем статистику по товарам
                    if item_name in product_stats:
                        product_stats[item_name] += item_quantity
                    else:
                        product_stats[item_name] = item_quantity
                
                items_text = "\n".join(items_text_parts)
                
                # Способ получения
                delivery_type = 'Доставка' if order['delivery_option'] == 'delivery' else 'Самовывоз'
                if delivery_type == 'Доставка':
                    summary_data['delivery_count'] += 1
                else:
                    summary_data['pickup_count'] += 1
                
                # Записываем данные
                ws1.cell(row=row_num, column=1, value=idx).font = cell_font
                ws1.cell(row=row_num, column=1).alignment = cell_alignment
                
                ws1.cell(row=row_num, column=2, value=order['id']).font = Font(name='Calibri', size=10, bold=True)
                ws1.cell(row=row_num, column=2).alignment = cell_alignment
                
                ws1.cell(row=row_num, column=3, value=order_time).font = cell_font
                ws1.cell(row=row_num, column=3).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                ws1.cell(row=row_num, column=4, value=order['user_name']).font = cell_font
                ws1.cell(row=row_num, column=4).alignment = cell_alignment
                
                ws1.cell(row=row_num, column=5, value=formatted_phone).font = cell_font
                ws1.cell(row=row_num, column=5).alignment = cell_alignment
                
                ws1.cell(row=row_num, column=6, value=items_text).font = product_font
                ws1.cell(row=row_num, column=6).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                ws1.cell(row=row_num, column=7, value=total_items_in_order).font = Font(name='Calibri', size=10, bold=True)
                ws1.cell(row=row_num, column=7).alignment = cell_alignment
                
                ws1.cell(row=row_num, column=8, value=order['total_amount']).font = money_font
                ws1.cell(row=row_num, column=8).alignment = cell_alignment
                ws1.cell(row=row_num, column=8).number_format = '#,##0 ₽'
                
                ws1.cell(row=row_num, column=9, value=order.get('discount_amount', 0)).font = Font(name='Calibri', size=10, color='E74C3C')
                ws1.cell(row=row_num, column=9).alignment = cell_alignment
                ws1.cell(row=row_num, column=9).number_format = '#,##0 ₽'
                
                ws1.cell(row=row_num, column=10, value=order['final_amount']).font = Font(name='Calibri', size=10, bold=True, color='27AE60')
                ws1.cell(row=row_num, column=10).alignment = cell_alignment
                ws1.cell(row=row_num, column=10).number_format = '#,##0 ₽'
                
                ws1.cell(row=row_num, column=11, value=delivery_type).font = cell_font
                ws1.cell(row=row_num, column=11).alignment = cell_alignment
                
                # Статус с цветом
                status_cell = ws1.cell(row=row_num, column=12, value=status_text)
                status_cell.font = status_font
                status_cell.alignment = cell_alignment
                status_cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                
                ws1.cell(row=row_num, column=13, value=order.get('comment', '')).font = Font(name='Calibri', size=9, color='7F8C8D')
                ws1.cell(row=row_num, column=13).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Добавляем границы
                for col_num in range(1, 14):
                    cell = ws1.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                
                # Подсчет итогов
                summary_data['total_amount'] += order['total_amount']
                summary_data['total_discount'] += order.get('discount_amount', 0)
                summary_data['total_final'] += order['final_amount']
                summary_data['total_products'] += total_items_in_order
                
                row_num += 1
            
            # Настраиваем ширину столбцов (пиксели → единицы openpyxl)
            column_widths = [5, 10, 12, 18, 23.57, 45, 12, 12, 12, 12, 10, 12, 25]
            for i, width in enumerate(column_widths, 1):
                ws1.column_dimensions[get_column_letter(i)].width = width
            
            # Автоматическая высота строк
            for row in range(5, row_num):
                items_cell = ws1.cell(row=row, column=6)  # Состав заказа
                lines = 1
                if items_cell.value:
                    lines = str(items_cell.value).count('\n') + 1
                ws1.row_dimensions[row].height = max(20, lines * 15)
            
            # Добавляем итоговую строку
            summary_row = row_num + 1
            
            # Объединяем ячейки для заголовка итогов
            ws1.merge_cells(f'A{summary_row}:F{summary_row}')
            ws1.cell(row=summary_row, column=1, value='📈 ИТОГОВАЯ СТАТИСТИКА').font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            ws1.cell(row=summary_row, column=1).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            ws1.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Итоговые показатели
            ws1.cell(row=summary_row, column=7, value=f"📦 {summary_data['total_products']} шт").font = Font(bold=True, color='2E86C1')
            ws1.cell(row=summary_row, column=7).alignment = cell_alignment
            
            ws1.cell(row=summary_row, column=8, value=f"💰 {summary_data['total_amount']:,} ₽").font = Font(bold=True, color='27AE60')
            ws1.cell(row=summary_row, column=8).alignment = cell_alignment
            ws1.cell(row=summary_row, column=8).number_format = '#,##0 ₽'
            
            ws1.cell(row=summary_row, column=10, value=f"💎 {summary_data['total_final']:,} ₽").font = Font(bold=True, color='E74C3C')
            ws1.cell(row=summary_row, column=10).alignment = cell_alignment
            ws1.cell(row=summary_row, column=10).number_format = '#,##0 ₽'
            
            # ===== ЛИСТ 2: СТАТИСТИКА ПО ТОВАРАМ =====
            ws2 = wb.create_sheet(title="📊 Топ товаров")
            
            # Заголовок
            ws2.merge_cells('A1:E1')
            ws2.cell(row=1, column=1, value="📦 Статистика продаж по товарам").font = title_font
            ws2.cell(row=1, column=1).fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
            ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Подзаголовок
            ws2.merge_cells('A2:E2')
            ws2.cell(row=2, column=1, value=f"Всего уникальных товаров: {len(product_stats)}").font = Font(name='Calibri', size=10, italic=True, color='7F7F7F')
            ws2.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Заголовки для статистики
            stats_headers = [
                '🏷️ Товар', '📦 Продано (шт)', '💰 Выручка (₽)', 
                '📊 Доля в продажах (%)', '🏅 Рейтинг'
            ]
            
            # Подсчитываем выручку по товарам
            product_revenue = {}
            for order in orders:
                try:
                    items = order['items']
                    if isinstance(items, str):
                        items = json.loads(items)
                    
                    for item in items:
                        item_name = item.get('name', 'Неизвестный товар')
                        item_total = item.get('total', 0)
                        
                        if item_name in product_revenue:
                            product_revenue[item_name] += item_total
                        else:
                            product_revenue[item_name] = item_total
                except:
                    pass
            
            # Сортируем товары по количеству продаж
            sorted_products = sorted(product_stats.items(), key=lambda x: x[1], reverse=True)
            
            for col_num, header in enumerate(stats_headers, 1):
                cell = ws2.cell(row=4, column=col_num, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Заполняем статистику
            stats_row = 5
            for idx, (product_name, quantity) in enumerate(sorted_products, 1):
                revenue = product_revenue.get(product_name, 0)
                percentage = (quantity / summary_data['total_products'] * 100) if summary_data['total_products'] > 0 else 0
                
                # Определяем рейтинг
                if idx == 1:
                    rating = "🥇 ЛИДЕР"
                elif idx == 2:
                    rating = "🥈 ТОП-2"
                elif idx == 3:
                    rating = "🥉 ТОП-3"
                elif idx <= 10:
                    rating = f"⭐ ТОП-{idx}"
                else:
                    rating = f"#{idx}"
                
                ws2.cell(row=stats_row, column=1, value=product_name).font = Font(name='Calibri', size=10, bold=True)
                ws2.cell(row=stats_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                
                ws2.cell(row=stats_row, column=2, value=quantity).font = Font(name='Calibri', size=10, bold=True, color='3498DB')
                ws2.cell(row=stats_row, column=2).alignment = cell_alignment
                
                ws2.cell(row=stats_row, column=3, value=revenue).font = Font(name='Calibri', size=10, bold=True, color='27AE60')
                ws2.cell(row=stats_row, column=3).alignment = cell_alignment
                ws2.cell(row=stats_row, column=3).number_format = '#,##0 ₽'
                
                ws2.cell(row=stats_row, column=4, value=round(percentage, 1)).font = Font(name='Calibri', size=10, color='8E44AD')
                ws2.cell(row=stats_row, column=4).alignment = cell_alignment
                ws2.cell(row=stats_row, column=4).number_format = '0.0"%"'
                
                ws2.cell(row=stats_row, column=5, value=rating).font = Font(name='Calibri', size=10, bold=True, color='E74C3C')
                ws2.cell(row=stats_row, column=5).alignment = cell_alignment
                
                # Добавляем границы
                for col_num in range(1, 6):
                    ws2.cell(row=stats_row, column=col_num).border = thin_border
                
                # Заливка для четных строк
                if idx % 2 == 0:
                    for col_num in range(1, 6):
                        ws2.cell(row=stats_row, column=col_num).fill = PatternFill(start_color='F8F9F9', end_color='F8F9F9', fill_type='solid')
                
                stats_row += 1
            
            # Настраиваем ширину столбцов
            stats_widths = [45, 15, 15, 15, 12]
            for i, width in enumerate(stats_widths, 1):
                ws2.column_dimensions[get_column_letter(i)].width = width
            
            # ===== ЛИСТ 3: АНАЛИТИКА И СВОДКА =====
            ws3 = wb.create_sheet(title="📈 Аналитика")
            
            # Заголовок
            ws3.merge_cells('A1:C1')
            ws3.cell(row=1, column=1, value="📊 Аналитическая сводка").font = title_font
            ws3.cell(row=1, column=1).fill = PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid')
            ws3.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Ключевые метрики
            metrics = [
                ("📊 Общее количество заказов", f"{summary_data['total_orders']:,}", "4F81BD"),
                ("📦 Всего товаров продано", f"{summary_data['total_products']:,} шт", "3498DB"),
                ("💰 Общая выручка", f"{summary_data['total_amount']:,} ₽", "27AE60"),
                ("🎫 Сумма скидок", f"{summary_data['total_discount']:,} ₽", "E74C3C"),
                ("💎 Итоговая сумма", f"{summary_data['total_final']:,} ₽", "9B59B6"),
                ("🚚 Заказов с доставкой", f"{summary_data['delivery_count']:,}", "F39C12"),
                ("🏪 Заказов самовывозом", f"{summary_data['pickup_count']:,}", "16A085"),
                ("📈 Средний чек", f"{round(summary_data['total_amount']/summary_data['total_orders'], 2):,} ₽", "2C3E50"),
            ]
            
            # Добавляем метрики
            metric_row = 3
            for i, (label, value, color) in enumerate(metrics):
                row = metric_row + (i // 2 * 2)
                col = (i % 2) * 3 + 1
                
                # Метка
                ws3.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
                label_cell = ws3.cell(row=row, column=col, value=label)
                label_cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                label_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Значение
                ws3.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
                value_cell = ws3.cell(row=row+1, column=col, value=value)
                value_cell.font = Font(name='Calibri', size=14, bold=True)
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = thin_border
            
            # Статистика по статусам
            status_row = metric_row + 10
            ws3.merge_cells(f'A{status_row}:C{status_row}')
            ws3.cell(row=status_row, column=1, value="📋 Распределение по статусам").font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            ws3.cell(row=status_row, column=1).fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            ws3.cell(row=status_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Добавляем статусы
            status_row += 1
            for status_id, (status_name, status_color) in status_names.items():
                count = summary_data['status_counts'].get(status_id, 0)
                percentage = (count / summary_data['total_orders'] * 100) if summary_data['total_orders'] > 0 else 0
                
                # Метка статуса
                ws3.cell(row=status_row, column=1, value=status_name).font = Font(name='Calibri', size=10, bold=True)
                ws3.cell(row=status_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                ws3.cell(row=status_row, column=1).fill = PatternFill(start_color=status_color + '20', end_color=status_color + '20', fill_type='solid')
                
                # Количество
                ws3.cell(row=status_row, column=2, value=count).font = Font(name='Calibri', size=10, bold=True)
                ws3.cell(row=status_row, column=2).alignment = cell_alignment
                
                # Процент
                ws3.cell(row=status_row, column=3, value=f"{percentage:.1f}%").font = Font(name='Calibri', size=10)
                ws3.cell(row=status_row, column=3).alignment = cell_alignment
                ws3.cell(row=status_row, column=3).number_format = '0.0"% "'
                
                status_row += 1
            
            # Настраиваем ширину столбцов
            ws3.column_dimensions['A'].width = 25
            ws3.column_dimensions['B'].width = 15
            ws3.column_dimensions['C'].width = 15
            ws3.column_dimensions['D'].width = 25
            ws3.column_dimensions['E'].width = 15
            ws3.column_dimensions['F'].width = 15
            
            # ===== ЛИСТ 4: ДЕТАЛЬНЫЙ ЧЕК =====
            ws4 = wb.create_sheet(title="🧾 Детальные чеки")
            
            # Заголовок
            ws4.merge_cells('A1:G1')
            ws4.cell(row=1, column=1, value="🧾 Подробные чеки по заказам").font = title_font
            ws4.cell(row=1, column=1).fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
            ws4.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Заголовки для детальных чеков
            check_headers = [
                'Заказ №', 'Дата', 'Клиент', 'Товар', 
                'Кол-во', 'Цена (₽)', 'Сумма (₽)'
            ]
            
            for col_num, header in enumerate(check_headers, 1):
                cell = ws4.cell(row=3, column=col_num, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Заполняем детальные чеки
            check_row = 4
            for order in orders:
                # Форматируем дату (МОСКОВСКОЕ ВРЕМЯ)
                order_time = ''
                if order.get('created_at'):
                    try:
                        # Преобразуем UTC время в московское
                        moscow_dt = self.convert_utc_to_moscow(order['created_at'])
                        if moscow_dt:
                            order_time = moscow_dt.strftime('%d.%m.%Y')
                        else:
                            order_time = str(order['created_at'])
                    except:
                        order_time = str(order['created_at'])
                
                # Получаем товары
                items = []
                try:
                    if isinstance(order['items'], str):
                        items = json.loads(order['items'])
                    else:
                        items = order['items']
                except:
                    items = []
                
                # Добавляем каждый товар
                for item in items:
                    ws4.cell(row=check_row, column=1, value=order['id']).font = Font(name='Calibri', size=10, bold=True)
                    ws4.cell(row=check_row, column=1).alignment = cell_alignment
                    
                    ws4.cell(row=check_row, column=2, value=order_time).font = cell_font
                    ws4.cell(row=check_row, column=2).alignment = cell_alignment
                    
                    ws4.cell(row=check_row, column=3, value=order['user_name']).font = cell_font
                    ws4.cell(row=check_row, column=3).alignment = cell_alignment
                    
                    ws4.cell(row=check_row, column=4, value=item.get('name', 'Неизвестный товар')).font = product_font
                    ws4.cell(row=check_row, column=4).alignment = Alignment(horizontal='left', vertical='center')
                    
                    ws4.cell(row=check_row, column=5, value=item.get('quantity', 0)).font = cell_font
                    ws4.cell(row=check_row, column=5).alignment = cell_alignment
                    
                    ws4.cell(row=check_row, column=6, value=item.get('price', 0)).font = money_font
                    ws4.cell(row=check_row, column=6).alignment = cell_alignment
                    ws4.cell(row=check_row, column=6).number_format = '#,##0 ₽'
                    
                    ws4.cell(row=check_row, column=7, value=item.get('total', 0)).font = Font(name='Calibri', size=10, bold=True, color='27AE60')
                    ws4.cell(row=check_row, column=7).alignment = cell_alignment
                    ws4.cell(row=check_row, column=7).number_format = '#,##0 ₽'
                    
                    # Добавляем границы
                    for col_num in range(1, 8):
                        ws4.cell(row=check_row, column=col_num).border = thin_border
                    
                    check_row += 1
            
            # Настраиваем ширину столбцов
            check_widths = [10, 12, 18, 40, 10, 12, 12]
            for i, width in enumerate(check_widths, 1):
                ws4.column_dimensions[get_column_letter(i)].width = width
            
            # Настраиваем страницы для печати
            for ws in [ws1, ws2, ws3, ws4]:
                ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.3, footer=0.3)
                ws.sheet_view.showGridLines = False
            
            # Сохраняем файл
            print("📁 Сохраняем Excel файл...")
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', mode='wb') as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name
                print(f"✅ Файл создан: {tmp_path}")
            
            try:
                print("📤 Отправляем файл в Telegram...")
                with open(tmp_path, 'rb') as f:
                    resp = requests.post(
                        f'https://api.telegram.org/bot{bot_token}/sendDocument',
                        data={
                            'chat_id': user_id, 
                            'caption': '📊 Профессиональный отчет АРТФЛОРА\n\n• 📋 Детализация заказов\n• 📊 Топ товаров\n• 📈 Аналитика\n• 🧾 Детальные чеки\n\nОтчет сформирован автоматически.'
                        },
                        files={'document': ('Отчет_АРТФЛОРА.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                        timeout=30
                    )
                
                print(f"📩 Ответ Telegram API: {resp.status_code}")
                
                if resp.status_code == 200:
                    print("✅ Excel файл успешно отправлен")
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    response_data = {'success': True, 'message': 'Отчет отправлен в Telegram'}
                    self.wfile.write(json.dumps(response_data).encode('utf-8'))
                else:
                    error_text = resp.text[:200] if resp.text else 'Неизвестная ошибка'
                    print(f"❌ Ошибка Telegram API: {resp.status_code} - {error_text}")
                    self.send_response(500)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    response_data = {'success': False, 'error': f'Ошибка отправки: {resp.status_code}'}
                    self.wfile.write(json.dumps(response_data).encode('utf-8'))
                    
            finally:
                # Удаляем временный файл
                try:
                    os.unlink(tmp_path)
                    print("🗑 Файл удален")
                except:
                    pass
                
        except Exception as e:
            print(f"💥 Ошибка при создании Excel: {e}")
            import traceback
            traceback.print_exc()
            # Пробуем создать CSV как fallback
            return self.export_to_csv(orders, bot_token, user_id)

    def export_to_csv(self, orders, bot_token, user_id):
        """Резервный метод для создания CSV файла"""
        try:
            print("📊 Создаем CSV файл (резервный метод)...")
            
            # Создаем CSV файл в памяти
            output = io.StringIO()
            csv_writer = csv.writer(output)
            
            # Заголовки CSV
            headers = ['ID', 'Дата и время', 'Клиент', 'Телефон', 'Сумма', 'Скидка', 'Итог', 
                      'Способ', 'Адрес', 'Статус', 'Комментарий']
            csv_writer.writerow(headers)
            
            status_names = {
                1: 'Новый',
                2: 'Подтвержден',
                3: 'Собирается',
                4: 'В пути',
                5: 'Доставлен',
                6: 'Отменен'
            }
            
            for order in orders:
                order_time = ''
                if order.get('created_at'):
                    try:
                        # Преобразуем UTC время в московское
                        moscow_dt = self.convert_utc_to_moscow(order['created_at'])
                        if moscow_dt:
                            order_time = moscow_dt.strftime('%d.%m.%Y %H:%M')
                        else:
                            order_time = order['created_at']
                    except:
                        order_time = order['created_at']
                
                row = [
                    order['id'],
                    order_time,
                    order['user_name'],
                    order['phone'],
                    order['total_amount'],
                    order.get('discount_amount', 0),
                    order['final_amount'],
                    'Доставка' if order['delivery_option'] == 'delivery' else 'Самовывоз',
                    order.get('delivery_address', ''),
                    status_names.get(order['status_id'], 'Новый'),
                    (order.get('comment', '')[:50] + '...') if len(order.get('comment', '')) > 50 else order.get('comment', '')
                ]
                csv_writer.writerow(row)
            
            # Создаем файл в памяти
            csv_data = output.getvalue().encode('utf-8')
            
            print("📁 Создаем временный файл CSV...")
            with tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='wb') as tmp:
                tmp.write(csv_data)
                tmp_path = tmp.name
                print(f"✅ Временный файл создан: {tmp_path}")
            
            try:
                print("📤 Отправляем CSV файл в Telegram...")
                with open(tmp_path, 'rb') as f:
                    resp = requests.post(
                        f'https://api.telegram.org/bot{bot_token}/sendDocument',
                        data={'chat_id': user_id, 'caption': '📊 Отчет по заказам в формате CSV'},
                        files={'document': ('orders_report.csv', f, 'text/csv')},
                        timeout=30
                    )
                
                print(f"📩 Ответ Telegram API: {resp.status_code}")
                
                if resp.status_code == 200:
                    print("✅ CSV файл успешно отправлен в Telegram")
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    response_data = {'success': True, 'message': 'CSV файл отправлен в Telegram'}
                    self.wfile.write(json.dumps(response_data).encode('utf-8'))
                else:
                    error_text = resp.text[:200] if resp.text else 'Неизвестная ошибка'
                    print(f"❌ Ошибка Telegram API: {resp.status_code} - {error_text}")
                    self.send_response(500)
                    self.send_header('Content-type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    response_data = {'success': False, 'error': f'Ошибка отправки файла: {resp.status_code}'}
                    self.wfile.write(json.dumps(response_data).encode('utf-8'))
                    
            finally:
                # Удаляем временный файл
                try:
                    os.unlink(tmp_path)
                    print("🗑 Временный файл удален")
                except:
                    pass
                
        except Exception as e:
            error_msg = str(e)
            print(f"💥 Ошибка при создании CSV: {error_msg}")
            
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            response_data = {'success': False, 'error': f'Ошибка создания файла: {error_msg}'}
            self.wfile.write(json.dumps(response_data).encode('utf-8'))

    def send_order_notification(self, order_id, status_id):
        try:
            bot_token = os.environ.get('BOT_TOKEN')
            
            if not bot_token:
                log_error("order_notification", "Missing BOT_TOKEN", "", "Failed to send order notification")
                return False
            
            order_response = supabase.table("orders").select("*").eq("id", order_id).execute()
            if not order_response.data:
                return False
            
            order = order_response.data[0]
            
            status_messages = {
                1: "✅ Ваш заказ принят! Мы начинаем его обработку.",
                2: "🔄 Заказ подтвержден! Мы готовим его к отправке.",
                3: "📦 Ваш заказ собирается! Скоро он будет у вас.",
                4: "🚗 Заказ в пути! Курьер уже везет его к вам.",
                5: "🎉 Заказ доставлен! Спасибо за покупку!",
                6: "❌ Заказ отменен."
            }
            
            message = status_messages.get(status_id, f"Статус заказа изменен")
            
            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
            payload = {
                'chat_id': order['user_id'],
                'text': message,
                'parse_mode': 'Markdown'
            }
            
            response = requests.post(url, json=payload, timeout=10)
            return response.status_code == 200
            
        except Exception as e:
            log_error("order_notification", e, "", f"Order ID: {order_id}")
            return False

    def update_promocode_usage(self, promocode_id):
        try:
            promocode_response = supabase.table("promocodes").select("used_count").eq("id", promocode_id).execute()
            if promocode_response.data:
                current_count = promocode_response.data[0].get('used_count', 0)
                supabase.table("promocodes").update({"used_count": current_count + 1}).eq("id", promocode_id).execute()
                
        except Exception as e:
            log_error("promocode_update", e, "", f"Promocode ID: {promocode_id}")
